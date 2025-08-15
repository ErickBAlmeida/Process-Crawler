import os
import re
import subprocess
import time

from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

class App:
    def __init__(self):
        
        load_dotenv()


        subprocess.Popen([
            os.getenv("CHROME_PATH"),
            f"--remote-debugging-port={os.getenv('DEBUG_PORT')}",
            f"--user-data-dir={os.getenv('USER_DATA_DIR')}",
            os.getenv("LINK_RJ")
        ])

        #   Configurações do Chrome para se conectar via DevTools
        options = Options()
        options.add_argument("--log-level=3")
        options.debugger_address = f"127.0.0.1:{os.getenv('DEBUG_PORT')}"        
        self.navegador = webdriver.Chrome(service=Service(), options=options)

    def logar(self):
        
        time.sleep(4)
        self.navegador.refresh()

        try:
            botao_menu = WebDriverWait(self.navegador, 15).until(
                EC.presence_of_element_located((By.CLASS_NAME, "botao-menu"))
            )
            botao_menu.click()
            print("\nNavegando pelo site...")

            time.sleep(1)
            self.navegador.find_element("partial link text", "Processo").click()

            time.sleep(.5)
            acoes = self.navegador.find_elements(By.XPATH, "//*[contains(text(), ' Outras ações ')]")	
            acoes[0].click()

            time.sleep(.5)
            soliciar_hab = self.navegador.find_elements(By.XPATH, "//*[contains(text(), ' Solicitar habilitação ')]")
            soliciar_hab[0].click()
            
        except Exception as e:

            time.sleep(5)
            raise Exception("Erro ao navegar pelo site") from e

    def ponteiro(self):
        wb = load_workbook("Bases\RJ.xlsx")
        sheet = wb.active
        
        cell_a = sheet.cell(row=2, column=1)
        num_processo = cell_a.value

        if num_processo is None:
            return None
        
        return num_processo

    def pesquisar(self, num_processo):
        
        try:
            processo_str = str(num_processo)

            print("="*50,'\n')
            print(f"Buscando por: {num_processo}\n\n")

            processo_str = re.sub(r'[^0-9]','', processo_str)

            if processo_str.isdigit() != True:
                print("\n\nPetição não é valida, indo para a proxima.\n\n")
                return False

            time.sleep(1)
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:numeroSequencial").send_keys(processo_str[0:7])
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:numeroDigitoVerificador").send_keys(processo_str[7:9])
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:Ano").send_keys(processo_str[9:13])
            self.navegador.find_element(By.ID, "fPP:numeroProcesso:NumeroOrgaoJustica").send_keys(processo_str[-4:])
            
            # Pesquisar Processo
            time.sleep(2)
            self.navegador.find_element(By.ID, "fPP:searchProcessos").click()

            time.sleep(1.5)

            try:
                self.navegador.find_element(By.CLASS_NAME, "rich-messages-label")
                print("O processo pesquisado é sigiloso!!\n\n")
                time.sleep(2)
                self.navegador.find_element("name", "fPP:clearButtonProcessos").click()

                self.res_polo = 'SIGILOSO'
                self.res_status = 'SIGILOSO'

                self.retorno(num_processo)                
                return False

            except:
                print("Abrindo processo...")
                time.sleep(2)

            btn_link = WebDriverWait(self.navegador, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div/div/div/div[2]/form/div[2]/div[2]/table/tbody/tr/td[2]/a'))
            )
            btn_link.click()

        except:
            
            time.sleep(3)
            print("❌ Erro ao pesquisar petição\n")
            raise

    def polo(self):        
        try:            
            time.sleep(3)
            abas = self.navegador.window_handles
            self.navegador.switch_to.window(abas[1])

            self.navegador.find_element(By.CLASS_NAME, "titulo-topo-desktop").click()
            time.sleep(1)

            polos_ativos = self.navegador.find_element(By.ID, "poloAtivo")

            if os.getenv('POLO_RJ') in polos_ativos.text:
                self.res_polo = 'Ativo'
                print("\n✅ POLO ATIVO")
                time.sleep(1)
                return True
        
        except:
            print("❌ Não foi possível localizar o polo ativo")
            self.res_polo = 'Inativo'
            self.res_status = 'Inativo'
            time.sleep(2)
            self.retorno()
            return False
        
    def status(self):
        status_map = {
            "arquivado": "Caso foi ARQUIVADO",
            "baixado": "Caso foi BAIXADO",
            "sentença": "Caso foi SENTENCIADO",
            "suspenso": "Caso foi SUSPENSO"
        }

        found = False
        page = self.navegador.page_source.lower()

        set_status = set()

        for termo, mensagem in status_map.items():
            if termo in page:
                set_status.add(termo)
                print(f"✅ {mensagem}")
                found = True

        list_status = list(set_status)

        if not found:
            print("🟨 NENHUM STATUS ENCONTRADO!!")
            self.res_status = 'N/D'
            time.sleep(1)

        else:
            self.res_status = ', '.join(str(x) for x in list_status)

    def retorno(self, num_processo):        
        try:
            retorno = [num_processo, self.res_polo, self.res_status]

            wb = load_workbook("Relatórios\saida_RJ.xlsx")
            sheet = wb.active

            sheet.append(retorno)
            wb.save("Relatórios\saida_RJ.xlsx")
            print("✅ Relatório atualizado com sucesso!!!")
        
        except Exception as e:
            print(f"Detalhes do erro: {e}")
            raise

    def atualizar_base(self):
            wb = load_workbook("Bases\RJ.xlsx")
            sheet = wb.active
            
            sheet.delete_rows(2, 1)
            wb.save("Bases\RJ.xlsx")
            print("✅ Base atualizada com sucesso.")

    def finalizar(self):
        abas = self.navegador.window_handles
        if len(abas) > 1:
            self.navegador.close()

        self.navegador.switch_to.window(abas[0])

        self.navegador.find_element(By.ID, "fPP:clearButtonProcessos").click()
        time.sleep(2)

    def run(self):
        self.logar()
        
        num_processo = self.ponteiro()
        
        if num_processo :
            if self.pesquisar(num_processo) != False:
                if self.polo() != False:
                    self.status()
                    
                self.retorno(num_processo)   
                self.atualizar_base()                 
                
                time.sleep(3)
                self.finalizar()
            
            else:
                print("Seguindo para o próximo processo...")  

if __name__ == "__main__":
    app = App()
    
    wb = load_workbook("Bases\RJ.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, max_col=1):
        app.run()