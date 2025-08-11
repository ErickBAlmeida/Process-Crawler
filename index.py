import os
import re
import subprocess
import time

from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from win10toast import ToastNotifier


class App:
    def __init__(self):
        
        load_dotenv()

        self.notifier = ToastNotifier()

        subprocess.Popen([
            os.getenv("CHROME_PATH"),
            f"--remote-debugging-port={os.getenv('DEBUG_PORT')}",
            f"--user-data-dir={os.getenv('USER_DATA_DIR')}",
            os.getenv("LINK")
        ])

        #   Configurações do Chrome para se conectar via DevTools
        options = Options()
        options.add_argument("--headless=new")
        options.debugger_address = f"127.0.0.1:{os.getenv('DEBUG_PORT')}"        
        self.navegador = webdriver.Chrome(service=Service(), options=options)

        self.run()

    def logar(self):
        self.navegador.find_element(By.ID, "identificacao").click()

        btnCertificado = WebDriverWait(self.navegador, 10).until(
            EC.presence_of_element_located((By.ID, "linkAbaCertificado"))
        )
        btnCertificado.click()

        time.sleep(3)
        self.navegador.find_element(By.ID, "submitCertificado").click()

    def navegar(self):
        burguerBtn = WebDriverWait(self.navegador, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "header__navbar__menu-hamburger"))
        )
        burguerBtn.click()

        time.sleep(1)
        self.navegador.find_element(By.XPATH, '//*[@id="root"]/div/header/nav/aside[1]/div[1]/nav/ul/li[2]/button').click()
        time.sleep(1)
        self.navegador.find_element(By.XPATH, '//*[@id="root"]/div/header/nav/aside[1]/div[1]/nav/ul/li[2]/ul/li[1]/a').click()

    def pesquisar(self, val):
        print(f"\nPesquisando processo: {val}")
        num_processo = re.sub(r'[^0-9]','', val)

        if num_processo.isdigit() != True:
            print("\n\n❌ Processo não é valida, indo para a proxima.\n\n")
            return False
        
        try:
            self.navegador.find_element(By.ID, 'numeroDigitoAnoUnificado').send_keys(num_processo[:13])
            self.navegador.find_element(By.XPATH, "//*[@id='foroNumeroUnificado']").send_keys(num_processo[-4:])
            self.navegador.find_element(By.ID, 'botaoConsultarProcessos').click()
        
        except:
            print("❌ Erro ao pesquisar o processo!")
        
    def ponteiro(self):
        wb = load_workbook("SP.xlsx")
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, max_col=1):
            cell_a = row[0]
            num_processo = str(cell_a.value).strip()
            self.linha = cell_a.row

            yield num_processo

    def polo(self):
        time.sleep(3)
        print("Buscando situação do Polo...")
        elemento = self.navegador.find_element(By.CLASS_NAME, "nomeParteEAdvogado").text
        
        try:
            if os.getenv("NOME_DO_POLO") in elemento.lower():
                print("✅ POLO ATIVO!!!")
                self.res_polo = "Ativo"
            
            else:
                print("❌ POLO INATIVO!!!")
                self.res_polo = "Inativo"
        
        except:
            print("❌ ERRO NA LOCALIZAÇÃO DO POLO!!!")
            raise

    def situProcesso(self):
        print("\nBuscando situação do processo...")

        labelSeg = None
        labelSitu = None

        try:
            labelSeg = self.navegador.find_element(By.ID, "labelSegredoDeJusticaProcesso")
            print(f"❌ O processo é um SEGREDO DE JUSTIÇA !!! \nSeguindo para o próximo...")
            self.res_situProcesso = "Segredo de Justiça"
            return False
        
        except:
            pass

        try:
            labelSitu = self.navegador.find_element(By.ID, "labelSituacaoProcesso")
            situ = labelSitu.text
            self.res_situProcesso = f"{situ.upper()}"
            print(f"❌ Processo {situ.upper()} !!! \nSeguindo para o próximo...")
            return False
        
        except:
            pass

        if labelSeg is None and labelSitu is None:
            print("✅ Processo em andamento!!")
            self.res_situProcesso = "Em Andamento"

    def status(self):
        print("\nBuscando status do processo...")

        status_map = {
            "arquivado": "Arquivado",
            "baixado": "Baixado",
            "sentença": "Sentença",
            "sentenciado": "Sentença",
        }

        link = WebDriverWait(self.navegador, 5).until(
            EC.presence_of_element_located((By.ID, "linkmovimentacoes"))
        )
        self.navegador.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", link)
        time.sleep(1)
        link.click()
        
        time.sleep(1)
        div_mov = self.navegador.find_element(By.ID, 'tabelaTodasMovimentacoes')
        mov_txt = div_mov.text

        var = False
        list_status = []

        # Busca status simples
        for termo, descricao in status_map.items():
            if termo.lower() in mov_txt.lower():
                list_status.append(descricao)
                print(f"✅ Caso está {descricao.upper()}")
                var = True
            
        if "Julgado" in mov_txt:
            print("✅ Caso está JULGADO")
            var = True

            if "Procedente" in mov_txt:
                list_status.append("Procedente")
                print("   ✅ Caso está JULGADO PROCEDENTE")
                var = True

            elif "improcedente" in mov_txt:
                list_status.append("Improcedente")
                print("   ✅ Caso está JULGADO IMPROCEDENTE")
                var = True

            else:
                list_status.append("Julgamento INDERTERMINADO")
                print("   🟨 Julgamento INDERTERMINADO!!!")


        if not var:
            self.res_status = ''
            print("🟨 NENHUM STATUS ENCONTRADO!!\n")
        
        else:
            self.res_status = ', '.join(str(x) for x in list_status)

        print() 

    def retorno(self, num_processo):
        
        try:
            retorno = [num_processo, self.res_polo, self.res_situProcesso, self.res_status]

            wb = load_workbook("saida_SP.xlsx")
            sheet = wb.active

            sheet.append(retorno)
            wb.save("saida_SP.xlsx")
        
        except:
            print("❌ Erro ao retornar arquivo Excel!!!")

        
    def run(self):
        self.logar()
        self.navegar()
        
        for num_processo in self.ponteiro():
            if self.pesquisar(num_processo) != False:
                self.polo()
                
                if self.situProcesso() != False:
                    self.status()
                    
                self.retorno(num_processo)                    
                
                time.sleep(3)
                print('='*50)
                self.navegador.find_element(By.ID, 'setaVoltar').click()
            
            else:
                print("Seguindo para o próximo processo...")
                continue       

try:
    app = App()

except Exception as e:
    print("\n\n❌ ERROR!!!\n\n")